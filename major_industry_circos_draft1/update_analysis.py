import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Configuration ---
INPUT_FILE = 'Ire Anh Data 1.22.26 (1).xlsx'
OUTPUT_FILE = 'Major_Career_Analysis_v4.xlsx'

# --- Data Loading & Cleaning ---

def clean_major_name(major):
    if pd.isna(major):
        return "Unknown"
    major = str(major).strip()
    
    # Remove degree suffixes like (BS), (BA)
    major = major.replace(" (BS)", "").replace(" (BA)", "")
    
    # Merge Computer Science variants
    if "Computer Science" in major and "Master Track" not in major:
        return "Computer Science"
    
    # Merge Environmental Science variants
    if major.startswith("Environmental Sci ("):
        return "Environmental Science"
    
    return major

def load_and_process_data():
    print(f"Loading data from {INPUT_FILE}...")
    df = pd.read_excel(INPUT_FILE)
    
    print(f"Initial rows: {len(df)}")
    
    # 1. Filter Degree Level
    # Include only Bachelor's
    # "Bachelor's Degree", "Bachelor's degree" -> standardize?
    # Actually, let's just exclude the ones we don't want.
    exclude_degrees = ['Graduate level degree', "Associate's degree", 'Master', 'MS', 'MBA']
    # But strictly, the user said "remove all the Master, MS, MBA, Associate degree".
    # And keep "Bachelor's"
    
    # Let's normalize Degree Level to check
    df['Degree Level Normalized'] = df['Degree Level'].astype(str).str.lower()
    
    # Filter
    df = df[~df['Degree Level Normalized'].str.contains('graduate|master|mba|associate', case=False, na=False)]
    print(f"Rows after Degree filter: {len(df)}")
    
    # 2. Filter Major - Remove Liberal Arts (professor's request)
    df = df[df['Program Name/Major'] != "Liberal Arts"]
    print(f"Rows after Liberal Arts filter: {len(df)}")
    
    # 3. Also filter out known graduate-only programs by name
    graduate_programs = [
        "MBA", "Executive MBA", "Health Care MBA",
        "MS Business Analytics", "MS Operations & Supply Chain Mgmt",
        "Comp Science BS (Master Track)",
        "Counseling Psychology", "Organization Develop & Change", "Organization Development",
        "Educ Leadership & Learning", "Educational Leadership & Admin",
        "Social Work Advanced Standing", "U.S. Law",
        "Leadership In Student Affairs", "Leadership", "Leadership & Management",
        "Technology Management", "Regulatory Science",
        "Publc Safety & Law Enfr Ldrshp", "Health Care Innovation", "Software Management"
    ]
    df = df[~df['Program Name/Major'].isin(graduate_programs)]
    print(f"Rows after graduate program filter: {len(df)}")
    
    # 3. Clean Major Names
    df['Cleaned Major'] = df['Program Name/Major'].apply(clean_major_name)
    
    return df

def clean_job_title(title):
    if pd.isna(title):
        return "Unknown"
    title = str(title).strip()
    original = title
    title_lower = title.lower()
    
    # Standardize Nursing
    if title_lower in ["rn", "registered nurse", "nurse"]:
        return "Registered Nurse"
        
    # Standardize common roles
    if "software engineer" in title_lower: return "Software Engineer"
    if "software developer" in title_lower: return "Software Developer" # Or merge with Engineer?
    if "project manager" in title_lower: return "Project Manager"
    if "account executive" in title_lower: return "Account Executive"
    if "teacher" in title_lower: return "Teacher"
    if "sales associate" in title_lower: return "Sales Associate"
    if "intern" in title_lower: return "Intern" # Maybe exclude interns? strict analysis might want full time. Keeping for now.
    
    return original

# --- Clustering Logic ---

    return original

# --- Clustering Logic ---

# 1. Job -> Small Cluster
def map_job_to_small_cluster(title):
    title = str(title).lower()
    
    # --- Business & Finance ---
    if any(k in title for k in ['accountant', 'audit', 'tax', 'bookkeep']): return "Accounting & Audit"
    if any(k in title for k in ['financ', 'wealth', 'bank', 'invest', 'risk', 'actuar', 'underwrit', 'loan', 'credit', 'equity', 'portfolio']): return "Finance & Investment"
    if any(k in title for k in ['project manager', 'program manager', 'operations', 'logistics', 'supply chain', 'coordinator', 'planner', 'production', 'inventory']): return "Management & Operations"
    if any(k in title for k in ['sales', 'account executive', 'market', 'brand', 'pr ', 'public relations', 'social media', 'advertising', 'digital', 'seo', 'communication', 'buyer', 'merchandis']): return "Sales & Marketing"

    # --- Technology & Engineering ---
    if any(k in title for k in ['software', 'developer', 'programmer', 'data', 'cloud', 'cyber', 'security', 'web', 'full stack', 'front end', 'back end']): return "Software & Data"
    if any(k in title for k in ['engineer', 'mechanical', 'civil', 'electrical', 'manufacturing', 'systems']):
        if "software" not in title: return "Engineering"
    if any(k in title for k in ['it ', 'network', 'technician', 'support', 'help desk', 'admin', 'tech ']): return "IT & Infrastructure"

    # --- Healthcare & Science ---
    if any(k in title for k in ['nurse', 'rn', 'doctor', 'physician', 'medical', 'patient', 'health', 'therapist', 'pharm', 'dent', 'care', 'hospital', 'clinic', 'veterinary']): return "Clinical Care"
    if any(k in title for k in ['research', 'scientist', 'bio', 'chem', 'lab', 'assistant', 'fellow']): 
        if "sales" not in title: return "Research & Science"
        
    # --- Education & Service ---
    if any(k in title for k in ['teach', 'tutor', 'professor', 'instructor', 'educat', 'school', 'faculty', 'academic', 'coach']): return "Education"
    if any(k in title for k in ['social work', 'counsel', 'case manager', 'advocate', 'community', 'youth', 'family', 'psych', 'nonprofit']): return "Social Service"
    
    # --- Arts, Media & Legal ---
    if any(k in title for k in ['design', 'writ', 'editor', 'artist', 'media', 'video', 'photog', 'journ', 'architect', 'interior', 'fashion']): return "Creative & Media"
    if any(k in title for k in ['legal', 'law', 'attorney', 'paralegal', 'compliance']): return "Legal & Policy"
    
    # --- Leadership ---
    if any(k in title for k in ['manager', 'director', 'executive', 'chief', 'president', 'lead', 'supervis', 'head of', 'owner', 'founder']): 
         # Catch-all for leadership if not caught above
         return "Leadership (General)"

    return "Unclassified Jobs"

# 2. Small Cluster -> Large Cluster
CAREER_SMALL_TO_LARGE = {
    # Business & Finance
    "Accounting & Audit": "Business & Finance",
    "Finance & Investment": "Business & Finance",
    "Management & Operations": "Business & Finance",
    "Sales & Marketing": "Business & Finance",
    
    # Technology & Engineering
    "Software & Data": "Technology & Engineering",
    "Engineering": "Technology & Engineering",
    "IT & Infrastructure": "Technology & Engineering",
    
    # Healthcare & Science
    "Clinical Care": "Healthcare & Science",
    "Research & Science": "Healthcare & Science",
    
    # Education & Service
    "Education": "Education & Service",
    "Social Service": "Education & Service",
    
    # Arts, Media & Legal
    "Creative & Media": "Arts, Media & Legal",
    "Legal & Policy": "Arts, Media & Legal",
    
    "Leadership (General)": "Business & Finance", # Defaulting general leadership to Business/Finance
    "Unclassified Jobs": "Unclassified"
}

def classify_career_hierarchy(title):
    small_cluster = map_job_to_small_cluster(title)
    large_cluster = CAREER_SMALL_TO_LARGE.get(small_cluster, "Unclassified")
    return small_cluster, large_cluster

def get_large_cluster(major):
    # Map from Small Cluster to Large Cluster directly to ensure consistency?
    # Or map Major -> Small -> Large
    pass

SMALL_CLUSTERS = {
    # Business & Management
    "Finance & Accounting": ["Accounting", "Actuarial Science", "Financial Management", "Risk Management and Insurance"],
    "General Business": ["Gen Business Mgmt", "Entrepreneurship", "Family Business", "Business Admin"],
    "Economics": ["Economics", "Economics - Business", "Economics - International", "Economics - Mathematical", "Economics - Public Policy"],
    "Marketing & Communication": ["Marketing Management", "Bus Admin - Communication", "Business Communication"],
    "Operations & Analytics": ["Operations Management", "Data Analytics"],
    "Specialized Business": ["International Business", "Real Estate Studies", "Human Resources Management", "Law & Compliance", "Org. Ethics & Compliance"],
    
    # Engineering & Technology
    "Computer Science & IT": ["Computer Science", "Information Technology", "Software Engineering", "Data Science", "Quant Methods - Computer Sci"],
    "Engineering Disciplines": ["Mechanical Engineering", "Electrical Engineering", "Civil Engineering", "Manufacturing Engineering", "Systems Engineering", "Computer Engineering"],
    
    # Natural & Health Sciences
    "Biological Sciences": ["Biology", "Biochemistry", "Neuroscience", "Biology of Global Health"],
    "Physical Sciences": ["Chemistry", "Physics", "Geology"],
    "Environmental Sciences": ["Environmental Science", "Environmental Studies", "Geography", "Geography - Geo Info Sys (GIS)"],
    "Health & Wellness": ["Nursing", "Exercise Science", "Public Health", "Health Promotion & Wellness"],
    
    # Social Sciences & Humanities
    "Social Sciences": ["Psychology", "Sociology", "Political Science", "Criminal Justice", "Justice & Peace Studies"],
    "Humanities": ["History", "Philosophy", "Art History", "Classical Civilization"],
    "International Studies": ["Intl Studies - Economics", "Intl Studies - History", "Intl Studies - Pol Sci"],
    
    # Communication & Media
    "Journalism & Media": ["Journalism", "COJO Journalism", "Digital Media Arts", "COJO Creative Multimedia"],
    "Communication Studies": ["Communication Studies", "Communication and Journalism", "COJO Interpersonal Comm", "COJO Persuasion/Soc Influence", "COJO Strategic Communications", "Strategic Comm: Ad and PR", "Strategic Communication"],
    "Creative Writing": ["English - Creative Writing", "Creative Writing & Publishing"],
    
    # Education & Social Services
    "Teacher Education": ["Elementary Education (K-6)", "Middle/Secondary Education", "K-12 Music Education", "K-12 World Lang. & Cultures", "Teacher Preparation - K-12", "Teacher Preparation-Elem K-6", "Teacher Preparation-Secondary", "Music Education"],
    "Educational Leadership": ["Educational Studies"],
    "Special Education & Counseling": ["Early Childhood Special Educ", "Autism Spectrum Disorders", "Acad Behavioral Strategist"],
    "Social Work": ["Social Work"],
    
    # Arts, Languages & Theology
    "Languages": ["French", "German", "Spanish", "Spanish Cultural/Literary St.", "Spanish Linguistics/Lang. St."],
    "English & Writing": ["English", "English - Professional Writing"],
    "Arts & Theology": ["Music", "Music - Business", "Music - Performance", "Theology", "Catholic Studies", "Pastoral Leadership", "Art History"]
}

LARGE_CLUSTERS = {
    "BUSINESS & MANAGEMENT": ["Finance & Accounting", "General Business", "Economics", "Marketing & Communication", "Operations & Analytics", "Specialized Business"],
    "ENGINEERING & TECHNOLOGY": ["Computer Science & IT", "Engineering Disciplines"],
    "NATURAL & HEALTH SCIENCES": ["Biological Sciences", "Physical Sciences", "Environmental Sciences", "Health & Wellness"],
    "SOCIAL SCIENCES & HUMANITIES": ["Economics", "Social Sciences", "Humanities", "International Studies"],
    "COMMUNICATION & MEDIA": ["Journalism & Media", "Communication Studies", "Creative Writing"],
    "EDUCATION & SOCIAL SERVICES": ["Teacher Education", "Educational Leadership", "Special Education & Counseling", "Social Work"],
    "ARTS, LANGUAGES & THEOLOGY": ["Languages", "English & Writing", "Arts & Theology"]
}

# Invert mappings for easy lookup
MAJOR_TO_SMALL = {}
for small, majors in SMALL_CLUSTERS.items():
    for m in majors:
        MAJOR_TO_SMALL[m] = small

SMALL_TO_LARGE = {}
for large, smalls in LARGE_CLUSTERS.items():
    for s in smalls:
        SMALL_TO_LARGE[s] = large

def get_cluster_info(major_name):
    # Try exact match first
    small = MAJOR_TO_SMALL.get(major_name)
    
    # If not found, try fuzzy match or defaults
    if not small:
        # Fallback logic
        if "Engineering" in major_name: small = "Engineering Disciplines"
        elif "Business" in major_name: small = "General Business"
        elif "Education" in major_name: small = "Teacher Education"
        elif "Music" in major_name: small = "Arts & Theology"
        elif "Comm" in major_name: small = "Communication Studies"
        elif "Bio" in major_name: small = "Biological Sciences"
        else: small = "Unclassified"
    
    large = SMALL_TO_LARGE.get(small, "Unclassified")
    return small, large


# --- Excel Generation ---

def generate_excel(df):
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. Overview
    ws = wb.create_sheet("Overview")
    ws.append(["Analysis Overview"])
    ws.append(["Total Students", len(df)])
    ws.append(["Unique Majors", df['Cleaned Major'].nunique()])
    ws['A1'].font = Font(size=14, bold=True)

    # 2. All Majors
    ws_majors = wb.create_sheet("All Majors")
    ws_majors.append(["Rank", "Major", "Count", "Percentage"])
    
    major_counts = df['Cleaned Major'].value_counts()
    for idx, (major, count) in enumerate(major_counts.items(), 1):
        pct = count / len(df)
        ws_majors.append([idx, major, count, f"{pct:.1%}"])
    
    # Style header
    for cell in ws_majors[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")

    # 3. Major Breakdown by Years (NEW)
    ws_years = wb.create_sheet("Breakdown by Years")
    
    # Pivot table
    # pivot = df.pivot_table(index='Cleaned Major', columns='Graduation Date', aggfunc='size', fill_value=0)
    # Pivot might fail if Graduation Date has bad values, let's clean it first?
    # Assuming Graduation Date is mainly years.
    
    pivot = pd.crosstab(df['Cleaned Major'], df['Graduation Date'])
    
    # Add columns to sheet
    cols = ["Major"] + list(pivot.columns)
    ws_years.append(cols)
    
    # Add rows
    for major in pivot.index:
        row = [major] + list(pivot.loc[major])
        ws_years.append(row)

    # Style header
    for cell in ws_years[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")

    # 4. Large Clusters
    ws_large = wb.create_sheet("Large Clusters")
    ws_large.append(["Large Cluster", "Count", "Percentage"])
    
    # Calculate counts
    large_counts = {}
    for major in df['Cleaned Major']:
        _, large = get_cluster_info(major)
        large_counts[large] = large_counts.get(large, 0) + 1
        
    for large, count in sorted(large_counts.items(), key=lambda x: x[1], reverse=True):
        ws_large.append([large, count, f"{count/len(df):.1%}"])

    # 5. Small Clusters
    ws_small = wb.create_sheet("Small Clusters")
    ws_small.append(["Small Cluster", "Large Cluster", "Count", "Percentage"])
    
    small_counts = {}
    small_to_large_map = {}
    for major in df['Cleaned Major']:
        small, large = get_cluster_info(major)
        small_counts[small] = small_counts.get(small, 0) + 1
        small_to_large_map[small] = large

    for small, count in sorted(small_counts.items(), key=lambda x: x[1], reverse=True):
        ws_small.append([small, small_to_large_map[small], count, f"{count/len(df):.1%}"])

    # 6. Cluster Breakdown
    ws_breakdown = wb.create_sheet("Cluster Breakdown")
    ws_breakdown.append(["Large Cluster", "Small Cluster", "Major", "Count"])
    
    # Create a list of tuples to sort
    breakdown_data = []
    for major, count in major_counts.items():
        small, large = get_cluster_info(major)
        breakdown_data.append((large, small, major, count))
    
    # Sort by Large, then Small, then Count desc
    breakdown_data.sort(key=lambda x: (x[0], x[1], -x[3]))
    
    for row in breakdown_data:
        ws_breakdown.append(row)

    # 7. Career Analysis (Job Titles)
    ws_career = wb.create_sheet("Career Analysis")
    
    # Add Cleaned Job Title and Cluster to DF for analysis
    df['Cleaned Job Title'] = df['Job Title'].apply(clean_job_title)
    
    # Apply hierarchy
    df[['Career Small Cluster', 'Career Large Cluster']] = df['Cleaned Job Title'].apply(
        lambda x: pd.Series(classify_career_hierarchy(x))
    )
    
    if 'Job Title' in df.columns:
        # Summary by Large Cluster
        ws_career.append(["Large Cluster Summary"])
        ws_career.append(["Large Cluster", "Count", "Percentage"])
        large_counts = df['Career Large Cluster'].value_counts()
        for cluster, count in large_counts.items():
             pct = count / len(df)
             ws_career.append([cluster, count, f"{pct:.1%}"])
             
        ws_career.append([])
        
        # Breakdown: Large -> Small -> Job (Top 50 Jobs for brevity?)
        # Let's list all small clusters and their top jobs
        ws_career.append(["Detailed Breakdown"])
        ws_career.append(["Large Cluster", "Small Cluster", "Job Title", "Count"])
        
        # Group by hierarchy
        breakdown = df.groupby(['Career Large Cluster', 'Career Small Cluster', 'Cleaned Job Title']).size().reset_index(name='Count')
        breakdown = breakdown.sort_values(['Career Large Cluster', 'Career Small Cluster', 'Count'], ascending=[True, True, False])
        
        for _, row in breakdown.iterrows():
            if row['Count'] > 1: # Only show jobs with > 1 student to keep list manageable? Or show all?
                # User asked for "fix the career spreadsheet to job title then small cluster, then big cluster"
                # This likely implies the structure of the sheet columns.
                ws_career.append([row['Career Large Cluster'], row['Career Small Cluster'], row['Cleaned Job Title'], row['Count']])

    # 8. Career Cluster Breakdown by Years (Small Cluster)
    ws_cluster_years = wb.create_sheet("Career Cluster by Years")
    pivot_cluster = pd.crosstab(df['Career Small Cluster'], df['Graduation Date'])
    
    # Add columns
    cols = ["Small Cluster"] + list(pivot_cluster.columns)
    ws_cluster_years.append(cols)
    for cluster in pivot_cluster.index:
        row = [cluster] + list(pivot_cluster.loc[cluster])
        ws_cluster_years.append(row)
        
    # Style
    for cell in ws_cluster_years[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD966", fill_type="solid")
        
    # 9. Job Breakdown by Years (NEW - Top 50)
    ws_job_years = wb.create_sheet("Job Titles by Years")
    
    # Filter for top job titles to avoid massive sparse sheet
    top_titles = df['Cleaned Job Title'].value_counts().head(50).index
    df_top_jobs = df[df['Cleaned Job Title'].isin(top_titles)]
    
    pivot_jobs = pd.crosstab(df_top_jobs['Cleaned Job Title'], df_top_jobs['Graduation Date'])
    
    cols = ["Job Title", "Small Cluster", "Large Cluster"] + list(pivot_jobs.columns)
    ws_job_years.append(cols)
    for job in pivot_jobs.index:
        # Lookup cluster info for this job
        small, large = classify_career_hierarchy(job)
        row = [job, small, large] + list(pivot_jobs.loc[job])
        ws_job_years.append(row)
        
    # Style
    for cell in ws_job_years[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="A9D08E", fill_type="solid")
    
    wb.save(OUTPUT_FILE)
    print(f"Analysis saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    df = load_and_process_data()
    generate_excel(df)
