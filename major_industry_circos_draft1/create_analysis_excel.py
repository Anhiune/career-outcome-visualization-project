import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the analysis results
with open('data_analysis_results.json', 'r') as f:
    data = json.load(f)

# Load original data for career analysis
df = pd.read_excel('Ire Anh Data 1.22.26 (1).xlsx')

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ===== SHEET 1: OVERVIEW =====
ws_overview = wb.create_sheet("Overview")
ws_overview.append(["Data Analysis Summary"])
ws_overview.append([])
ws_overview.append(["Total Students:", data['total_students']])
ws_overview.append(["Total Unique Majors:", len(data['majors'])])
ws_overview.append(["Career Data Columns:", ", ".join(data['career_columns'])])
ws_overview.append([])
ws_overview.append(["This workbook contains:"])
ws_overview.append(["- All Majors: Complete list with student counts"])
ws_overview.append(["- Large Clusters: 7 broad disciplinary categories"])
ws_overview.append(["- Small Clusters: 23 specific academic groupings"])
ws_overview.append(["- Career Analysis: Job titles and employers"])
ws_overview.append(["- Summary Stats: Distribution insights"])

# Style the overview
ws_overview['A1'].font = Font(size=16, bold=True)
ws_overview['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_overview['A1'].font = Font(size=16, bold=True, color="FFFFFF")

# ===== SHEET 2: ALL MAJORS =====
ws_majors = wb.create_sheet("All Majors")
ws_majors.append(["Rank", "Major Name", "Student Count", "Percentage"])

majors_df = pd.DataFrame([
    {"Major": major, "Count": data['major_counts'][major]}
    for major in data['majors']
]).sort_values('Count', ascending=False).reset_index(drop=True)

majors_df['Percentage'] = (majors_df['Count'] / data['total_students'] * 100).round(2)

for idx, row in majors_df.iterrows():
    ws_majors.append([idx + 1, row['Major'], row['Count'], f"{row['Percentage']}%"])

# Style headers
for cell in ws_majors[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# ===== SHEET 3: LARGE CLUSTERS =====
ws_large = wb.create_sheet("Large Clusters (7)")

# Define large clusters
large_clusters = {
    "BUSINESS & MANAGEMENT": [
        "Accounting", "Actuarial Science", "Financial Management", "Risk Management and Insurance",
        "Gen Business Mgmt", "Entrepreneurship", "Family Business",
        "Marketing Management", "Bus Admin - Communication", "Business Communication",
        "Operations Management", "Data Analytics",
        "International Business", "Real Estate Studies", "Human Resources Management", "Law & Compliance", "Org. Ethics & Compliance"
    ],
    "ENGINEERING & TECHNOLOGY": [
        "Computer Science", "Computer Engineering",
        "Software Engineering", "Data Science", "Information Technology", "Quant Methods - Computer Sci",
        "Mechanical Engineering", "Electrical Engineering", "Civil Engineering", "Manufacturing Engineering", "Systems Engineering"
    ],
    "NATURAL & HEALTH SCIENCES": [
        "Biology", "Biochemistry", "Neuroscience", "Biology of Global Health",
        "Chemistry", "Physics", "Geology",
        "Environmental Science", "Environmental Studies", "Geography", "Geography - Geo Info Sys (GIS)",
        "Nursing", "Exercise Science", "Public Health", "Health Promotion & Wellness"
    ],
    "SOCIAL SCIENCES & HUMANITIES": [
        "Psychology", "Sociology", "Political Science",
        "Economics", "Economics - Business", "Economics - International", "Economics - Mathematical", "Economics - Public Policy",
        "Criminal Justice", "Justice & Peace Studies", "History", "Philosophy",
        "Art History", "Classical Civilization", "Intl Studies - Economics", "Intl Studies - History", "Intl Studies - Pol Sci"
    ],
    "COMMUNICATION & MEDIA": [
        "Journalism", "COJO Journalism", "Digital Media Arts", "COJO Creative Multimedia",
        "Communication Studies", "Communication and Journalism", "COJO Interpersonal Comm",
        "COJO Persuasion/Soc Influence", "COJO Strategic Communications", "Strategic Comm: Ad and PR",
        "Strategic Communication", "English - Creative Writing", "Creative Writing & Publishing"
    ],
    "EDUCATION & SOCIAL SERVICES": [
        "Elementary Education (K-6)", "Middle/Secondary Education", "K-12 Music Education",
        "K-12 World Lang. & Cultures", "Teacher Preparation - K-12", "Teacher Preparation-Elem K-6",
        "Teacher Preparation-Secondary", "Music Education", "Educational Studies",
        "Early Childhood Special Educ", "Autism Spectrum Disorders",
        "Acad Behavioral Strategist", "Social Work"
    ],
    "ARTS, LANGUAGES & THEOLOGY": [
        "French", "German", "Spanish", "Spanish Cultural/Literary St.", "Spanish Linguistics/Lang. St.",
        "English", "English - Professional Writing", "Music", "Music - Business", "Music - Performance",
        "Mathematics (Applied Track)", "Mathematics (Education Track)", "Mathematics (Pure Track)",
        "Mathematics (Statistics Track)", "Statistics", "Theology", "Catholic Studies", "Pastoral Leadership",
        "Pastoral Ministry",
        "Legal Studies",
        "Individualized", "Family Studies"
    ]
}

ws_large.append(["Large Cluster", "Total Students", "Percentage", "Number of Majors"])

cluster_summary = []
for cluster_name, majors_list in large_clusters.items():
    total = sum(data['major_counts'].get(major, 0) for major in majors_list)
    pct = round(total / data['total_students'] * 100, 2)
    num_majors = len([m for m in majors_list if m in data['majors']])
    cluster_summary.append({
        'Cluster': cluster_name,
        'Total': total,
        'Percentage': pct,
        'Majors': num_majors
    })
    ws_large.append([cluster_name, total, f"{pct}%", num_majors])

# Style headers
for cell in ws_large[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# ===== SHEET 4: SMALL CLUSTERS =====
ws_small = wb.create_sheet("Small Clusters (23)")

small_clusters = {
    # Business & Management
    "Finance & Accounting": ["Accounting", "Actuarial Science", "Financial Management", "Risk Management and Insurance"],
    "General Business": ["Gen Business Mgmt", "Entrepreneurship", "Family Business"],
    "Marketing & Communication": ["Marketing Management", "Bus Admin - Communication", "Business Communication"],
    "Operations & Analytics": ["Operations Management", "Data Analytics"],
    "Specialized Business": ["International Business", "Real Estate Studies", "Human Resources Management", "Law & Compliance", "Org. Ethics & Compliance"],
    
    # Engineering & Technology
    "Computer Science & IT": ["Computer Science", "Software Engineering", "Data Science", "Information Technology", "Quant Methods - Computer Sci"],
    "Engineering Disciplines": ["Mechanical Engineering", "Electrical Engineering", "Civil Engineering", "Manufacturing Engineering", "Systems Engineering", "Computer Engineering"],
    
    # Natural & Health Sciences
    "Biological Sciences": ["Biology", "Biochemistry", "Neuroscience", "Biology of Global Health"],
    "Physical Sciences": ["Chemistry", "Physics", "Geology"],
    "Environmental Sciences": ["Environmental Science", "Environmental Studies", "Geography", "Geography - Geo Info Sys (GIS)"],
    "Health & Wellness": ["Nursing", "Exercise Science", "Public Health", "Health Promotion & Wellness"],
    
    # Social Sciences & Humanities
    "Economics": ["Economics", "Economics - Business", "Economics - International", "Economics - Mathematical", "Economics - Public Policy"],
    "Social Sciences": ["Psychology", "Sociology", "Political Science", "Criminal Justice", "Justice & Peace Studies"],
    "Humanities": ["History", "Philosophy", "Art History", "Classical Civilization"],
    "International Studies": ["Intl Studies - Economics", "Intl Studies - History", "Intl Studies - Pol Sci"],
    
    # Communication & Media
    "Journalism & Media": ["Journalism", "COJO Journalism", "Digital Media Arts", "COJO Creative Multimedia"],
    "Communication Studies": ["Communication Studies", "Communication and Journalism", "COJO Interpersonal Comm", "COJO Persuasion/Soc Influence", "COJO Strategic Communications", "Strategic Comm: Ad and PR", "Strategic Communication"],
    "Creative Writing": ["English - Creative Writing", "Creative Writing & Publishing"],
    
    # Education & Social Services
    "Teacher Education": ["Elementary Education (K-6)", "Middle/Secondary Education", "K-12 Music Education", "K-12 World Lang. & Cultures", "Teacher Preparation - K-12", "Teacher Preparation-Elem K-6", "Teacher Preparation-Secondary", "Music Education"],
    "Educational Leadership": ["Educational Studies"],
    "Special Education & Counseling": ["Early Childhood Special Educ", "Autism Spectrum Disorders", "Acad Behavioral Strategist"],
    "Social Work": ["Social Work"],
    
    # Arts, Languages & Theology
    "Languages": ["French", "German", "Spanish", "Spanish Cultural/Literary St.", "Spanish Linguistics/Lang. St."],
    "English & Writing": ["English", "English - Professional Writing"],
}

ws_small.append(["Small Cluster", "Large Cluster", "Total Students", "Percentage", "Number of Majors"])

for cluster_name, majors_list in small_clusters.items():
    # Find parent large cluster
    parent = None
    for large_name, large_majors in large_clusters.items():
        if any(m in large_majors for m in majors_list):
            parent = large_name
            break
    
    total = sum(data['major_counts'].get(major, 0) for major in majors_list)
    pct = round(total / data['total_students'] * 100, 2)
    num_majors = len([m for m in majors_list if m in data['majors']])
    ws_small.append([cluster_name, parent, total, f"{pct}%", num_majors])

# Style headers
for cell in ws_small[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


# ===== SHEET 4.5: CLUSTER BREAKDOWN =====
ws_breakdown = wb.create_sheet("Cluster Breakdown")
ws_breakdown.append(["Large Cluster", "Small Cluster", "Major Name", "Student Count", "% of Cluster"])

for cluster_name, majors_list in small_clusters.items():
    # Find parent large cluster
    parent = None
    for large_name, large_majors in large_clusters.items():
        if any(m in large_majors for m in majors_list):
            parent = large_name
            break
            
    # Calculate cluster total for percentage
    cluster_total = sum(data['major_counts'].get(major, 0) for major in majors_list)
    
    # Sort majors by count within the cluster
    sorted_majors = sorted(majors_list, key=lambda m: data['major_counts'].get(m, 0), reverse=True)
    
    for major in sorted_majors:
        if major in data['majors']:
            count = data['major_counts'].get(major, 0)
            if cluster_total > 0:
                pct = f"{round(count / cluster_total * 100, 1)}%"
            else:
                pct = "0%"
            ws_breakdown.append([parent, cluster_name, major, count, pct])

# Style headers
for cell in ws_breakdown[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

# ===== SHEET 5: CAREER ANALYSIS =====

ws_career = wb.create_sheet("Career Analysis")

# Analyze job titles
job_titles = df['Job Title'].dropna()
employers = df['Employing Organization'].dropna()

ws_career.append(["CAREER DATA SUMMARY"])
ws_career.append([])
ws_career.append(["Total Job Title Entries:", len(job_titles)])
ws_career.append(["Unique Job Titles:", job_titles.nunique()])
ws_career.append(["Total Employer Entries:", len(employers)])
ws_career.append(["Unique Employers:", employers.nunique()])
ws_career.append([])
ws_career.append(["TOP 30 JOB TITLES"])
ws_career.append(["Rank", "Job Title", "Count", "Percentage"])

top_jobs = job_titles.value_counts().head(30)
for idx, (job, count) in enumerate(top_jobs.items(), 1):
    pct = round(count / len(job_titles) * 100, 2)
    ws_career.append([idx, job, count, f"{pct}%"])

ws_career.append([])
ws_career.append(["TOP 30 EMPLOYERS"])
ws_career.append(["Rank", "Employer", "Count", "Percentage"])

top_employers = employers.value_counts().head(30)
for idx, (emp, count) in enumerate(top_employers.items(), 1):
    pct = round(count / len(employers) * 100, 2)
    ws_career.append([idx, emp, count, f"{pct}%"])

# Style
ws_career['A1'].font = Font(size=14, bold=True)
ws_career['A8'].font = Font(size=12, bold=True)

# ===== SHEET 6: SUMMARY STATISTICS =====
ws_stats = wb.create_sheet("Summary Statistics")

ws_stats.append(["DISTRIBUTION INSIGHTS"])
ws_stats.append([])
ws_stats.append(["Major Distribution"])
ws_stats.append(["Metric", "Value"])
ws_stats.append(["Average students per major", round(data['total_students'] / len(data['majors']), 2)])
ws_stats.append(["Median students per major", majors_df['Count'].median()])
ws_stats.append(["Most popular major", majors_df.iloc[0]['Major']])
ws_stats.append(["Students in most popular major", majors_df.iloc[0]['Count']])
ws_stats.append(["Least popular major", majors_df.iloc[-1]['Major']])
ws_stats.append(["Students in least popular major", majors_df.iloc[-1]['Count']])
ws_stats.append([])
ws_stats.append(["Large Cluster Distribution"])
ws_stats.append(["Largest cluster", max(cluster_summary, key=lambda x: x['Total'])['Cluster']])
ws_stats.append(["Students in largest cluster", max(cluster_summary, key=lambda x: x['Total'])['Total']])
ws_stats.append(["Smallest cluster", min(cluster_summary, key=lambda x: x['Total'])['Cluster']])
ws_stats.append(["Students in smallest cluster", min(cluster_summary, key=lambda x: x['Total'])['Total']])


# ===== JOB TITLE CLASSIFICATION =====
def classify_job_title(title):
    title = str(title).lower()
    
    # 1. Technical & Engineering
    if any(k in title for k in ['engineer', 'developer', 'software', 'data', 'analyst', 'programmer', 'technician', 'systems', 'network', 'it ', 'support', 'technical', 'cloud', 'cyber', 'web']):
        if not any(k in title for k in ['financial', 'business', 'marketing', 'sales', 'hr', 'recruiter']):
             return "Technical & Engineering"

    # 2. Finance & Analysis
    if any(k in title for k in ['financ', 'accountant', 'audit', 'tax', 'wealth', 'bank', 'invest', 'risk', 'actuar', 'underwrit', 'loan', 'credit']):
        return "Finance & Analysis"
        
    # 3. Marketing & Sales
    if any(k in title for k in ['market', 'sales', 'representative', 'account executive', 'brand', 'pr ', 'public relations', 'social media', 'advertising', 'digital', 'content', 'seo', 'communication']):
        return "Marketing & Sales"
        
    # 4. Operations & Logistics
    if any(k in title for k in ['operation', 'logistics', 'supply chain', 'project manager', 'program manager', 'coordinator', 'planner', 'scheduler', 'production', 'quality', 'inventory']):
        return "Operations & Logistics"

    # 5. Healthcare & Clinical
    if any(k in title for k in ['nurse', 'rn', 'doctor', 'physician', 'clinical', 'medical', 'patient', 'health', 'therapist', 'pharm', 'dent', 'care', 'hospital', 'clinic']):
        if not any(k in title for k in ['financial', 'sales', 'recruiter', 'hr']):
             return "Healthcare & Clinical"

    # 6. Education & Instruction
    if any(k in title for k in ['teach', 'tutor', 'professor', 'instructor', 'educat', 'school', 'faculty', 'academic', 'coach', 'trainer']):
        return "Education & Instruction"
        
    # 7. Research & Scientific
    if any(k in title for k in ['research', 'scientist', 'bio', 'chem', 'lab', 'assistant', 'fellow']):
        if not any(k in title for k in ['sales', 'market']):
            return "Research & Scientific"
            
    # 8. Counseling & Social Support
    if any(k in title for k in ['social work', 'counsel', 'case manager', 'advocate', 'community', 'youth', 'family', 'psych']):
        return "Counseling & Social Support"
        
    # 9. Creative & Media
    if any(k in title for k in ['design', 'writ', 'editor', 'artist', 'media', 'video', 'photog', 'journ', 'architect', 'interior', 'fashion']):
        return "Creative & Media"
        
    # 10. Leadership & Executive (and HR/Management)
    if any(k in title for k in ['manager', 'director', 'executive', 'chief', 'president', 'lead', 'supervis', 'human resource', 'hr ', 'recruit', 'talent']):
        return "Leadership & Management"
        
    return "Other / Unclassified"

# Get unique job titles
unique_titles = df['Job Title'].dropna().unique()

# Classify and count
job_classifications = {}
for title in unique_titles:
    cluster = classify_job_title(title)
    if cluster not in job_classifications:
        job_classifications[cluster] = []
    
    # Get count for this title
    count = len(df[df['Job Title'] == title])
    job_classifications[cluster].append((title, count))

# Create Sheet
ws_jobs = wb.create_sheet("Job Title Breakdown")
ws_jobs.append(["Career Cluster", "Job Title", "Count", "% of Cluster"])

for cluster, jobs in job_classifications.items():
    # Sort jobs by count
    sorted_jobs = sorted(jobs, key=lambda x: x[1], reverse=True)
    cluster_total = sum(count for _, count in sorted_jobs)
    
    for title, count in sorted_jobs:
        pct = round(count / cluster_total * 100, 1) if cluster_total > 0 else 0
        ws_jobs.append([cluster, title, count, f"{pct}%"])

# Style headers
for cell in ws_jobs[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

# Save workbook
output_file = 'Major_Career_Analysis_v2.xlsx'
wb.save(output_file)
print(f"✓ Excel workbook created: {output_file}")
print(f"  - {len(wb.sheetnames)} sheets")
print(f"  - {len(data['majors'])} majors analyzed")
print(f"  - {data['total_students']} total students")
