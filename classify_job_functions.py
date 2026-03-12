"""
Classify job titles into functional groups based on BOTH the job title AND company context.
This creates a 'Job Title Classified Group' that reflects what the person actually DOES,
not just what industry the company is in.
"""
import openpyxl
import re
from collections import Counter, defaultdict

INPUT_PATH = r'major_industry_circos_draft1\data\Career_Company_Industry_CLASSIFIED_updated.xlsx'
OUTPUT_PATH = r'major_industry_circos_draft1\data\Career_Company_Industry_CLASSIFIED_v2.xlsx'

# ── Job Function Categories ──────────────────────────────────────────────────
# Order matters: rules are checked top-to-bottom; first match wins.
# Each rule: (category_name, title_patterns, company_patterns_optional)
# Patterns are checked with re.search (case-insensitive).

CLASSIFICATION_RULES = [
    # ── 1. Education & Teaching ──────────────────────────────────────────────
    ("Education & Teaching", [
        r'\bteacher\b', r'\bteaching\b', r'\binstructor\b', r'\bprofessor\b',
        r'\bfaculty\b', r'\blecturer\b', r'\btutor\b', r'\beducator\b',
        r'\bparaprofessional\b', r'\bcurricul', r'\bgrade teacher\b',
        r'\bspecial education\b', r'\bspec\.?\s*ed\b', r'\bsped\b',
        r'\beach\b.*\bassist', r'\bteach\b',
        r'\badmissions\s*(counselor|recruiter)\b',
        r'\bacademic\s*(advisor|counselor|coordinator|program)\b',
        r'\bschool\s*(counselor|psycholog|social\s*worker)\b',
        r'\bresidence\s*(director|life)\b',
        r'\bstudent\s*(life|affairs|development|success|engagement)\b',
    ], None),

    # ── 2. Healthcare & Clinical ─────────────────────────────────────────────
    ("Healthcare & Clinical", [
        r'\bnurse\b', r'\bnursing\b', r'\bregistered\s*nurse\b', r'\brn\b',
        r'\bphysician\b', r'\bdoctor\b', r'\bmd\b', r'\bsurgeon\b',
        r'\bpharmac', r'\bdentist\b', r'\bdental\b',
        r'\btherapist\b', r'\btherapy\b', r'\bclinical\b', r'\bclinician\b',
        r'\bpatient\b', r'\bmedical\b',
        r'\bmental\s*health\b', r'\bbehavioral\s*(health|therapist|specialist|analyst)\b',
        r'\bcounselor\b', r'\bcounseling\b', r'\bpsycholog',
        r'\bsocial\s*worker\b', r'\blcsw\b', r'\blpc\b', r'\blmft\b',
        r'\bscribe\b', r'\bhealth\s*(care|coach|educator|navigator|specialist|aide)\b',
        r'\boccupational\b', r'\bphysical\s*therap', r'\bspeech',
        r'\bdiagnostic\b', r'\bradiolog', r'\bsonograph',
        r'\bdirect\s*(care|support)\b', r'\bcare\s*(aide|giver|coordinator|manager|specialist|worker|provider|assistant)\b',
        r'\bpara\s*medic\b', r'\bemt\b', r'\bchiropract',
        r'\boptom', r'\baudiol', r'\bortho', r'\bdermat',
        r'\bveterinar', r'\bvet\s*tech',
        r'\brespiratory\b', r'\bphlebotom', r'\bsurgical\s*tech',
        r'\baddiction\b', r'\bsubstance\s*abuse\b',
        r'\bhome\s*health\b', r'\bhospice\b',
        r'\bpublic\s*health\b',
        r'\bcna\b', r'\bmedical\s*assist',
        r'\bnursery\s*assist',  # child care nursing context
        r'\bcase\s*manage', r'\bcase\s*worker',
        r'\bambulatory\b',
        r'\bed\s*nurse\b',
        r'\brehab', 
    ], None),

    # ── 3. Software & IT ─────────────────────────────────────────────────────
    ("Software & IT", [
        r'\bsoftware\b', r'\bdeveloper\b', r'\bprogrammer\b',
        r'\bfull\s*stack\b', r'\bfront\s*end\b', r'\bback\s*end\b',
        r'\bweb\s*develop', r'\bmobile\s*develop',
        r'\bdevops\b', r'\bsre\b', r'\bsite\s*reliability\b',
        r'\bit\s+(analyst|manager|specialist|admin|coord|director|support|tech|engineer|consult|architect|intern)\b',
        r'\binformation\s*technology\b', r'\binformation\s*systems\b',
        r'\bsystems?\s*(admin|engineer|analyst|architect)\b',
        r'\bnetwork\s*(engineer|admin|analyst|architect)\b',
        r'\bcloud\b', r'\baws\b(?!\s*(application|app)\s*lead)',
        r'\bcyber\s*security\b', r'\binfosec\b', r'\bsecurity\s*(engineer|analyst|architect|specialist)\b',
        r'\bdatabase\b', r'\bdba\b', r'\bsql\b',
        r'\btechnology\s*(services|analyst|specialist|consultant)\b',
        r'\bui\s*/?\s*ux\b', r'\bux\s*(design|research|engineer)\b',
        r'\betl\b', r'\bplatform\s*engineer\b',
        r'\bqa\s*(engineer|analyst|tester|lead)\b',
        r'\btest\s*(engineer|automation)\b',
        r'\bhelp\s*desk\b', r'\btechnical\s*support\b',
        r'\bapplication\s*(develop|engineer|architect|lead|support)\b',
        r'\bsap\s', r'\bsalesforce\b', r'\bworkday\b',
        r'\bscrum\b', r'\bagile\b',
        r'\btech\s*(aid|lead|intern)\b',
        r'\bdesign\s*systems?\s*engineer\b',
        r'\bsubject\s*matter\s*expert\b',
    ], None),

    # ── 4. Data & Analytics ──────────────────────────────────────────────────
    ("Data & Analytics", [
        r'\bdata\s*(scien|analy|engineer|architect|govern|visual|manag|quality|special|strat|ops)\b',
        r'\bdata\s*$',  # just "Data" in title
        r'\bbusiness\s*intelligen', r'\bbi\s+(analyst|developer|engineer|manager)\b',
        r'\bmachine\s*learning\b', r'\bartificial\s*intelligen', r'\bai\s+engineer\b',
        r'\banalytics\s*(engineer|consult|manager|lead|special|ldp|analyst)\b',
        r'\bstatistic', r'\bquantitative\s*(analyst|research|developer)\b',
        r'\bmaster\s*data\b',
    ], None),

    # ── 5. Finance & Accounting ──────────────────────────────────────────────
    ("Finance & Accounting", [
        r'\bfinance\b', r'\bfinancial\b', r'\baccountant\b', r'\baccounting\b',
        r'\baudit\b', r'\bauditor\b', r'\btax\b', r'\btaxes\b',
        r'\btreasur', r'\bcontroller\b', r'\bcomptroller\b',
        r'\bcredit\s*(analyst|manager|officer|specialist)\b',
        r'\bloan\b', r'\bmortgage\b',
        r'\binvestment\b', r'\bportfolio\b', r'\bwealth\s*manag',
        r'\basset\s*manag', r'\bequity\b',
        r'\bactuar', r'\bunderwrite', r'\bunderwriting\b',
        r'\bfp&a\b', r'\bfinancial\s*plan',
        r'\bcfo\b', r'\bbook\s*keep', r'\bbookkeep',
        r'\baccounts\s*(receiv|payab)', r'\bcollection',
        r'\bclaim', r'\binsurance\s*(agent|analyst|specialist|underwriter|adjuster|broker)\b',
        r'\brisk\s*(analyst|manager|officer|consult)\b',
        r'\bcompliance\s*(analyst|officer|specialist|manager|assoc|coordinator)\b',
        r'\bvaluation\b', r'\bappraiser\b', r'\bappraisal\b',
        r'\bbanker\b', r'\bbanking\b', r'\bteller\b',
        r'\bpayroll\b',
    ], None),

    # ── 6. Actuarial Science ─────────────────────────────────────────────────
    # (Already caught by Finance above via r'\bactuar')

    # ── 7. Engineering (Non-Software) ────────────────────────────────────────
    ("Engineering", [
        r'\bengineer\b', r'\bengineering\b', r'\bwngineer\b',  # catch typo
        r'\bmechanical\b', r'\belectrical\b', r'\bcivil\b',
        r'\bchemical\s*engineer\b', r'\bindustrial\s*engineer\b',
        r'\bmanufacturing\s*engineer\b', r'\bprocess\s*engineer\b',
        r'\bquality\s*engineer\b', r'\bdesign\s*engineer\b',
        r'\bproject\s*engineer\b', r'\bfield\s*engineer\b',
        r'\bstructural\b', r'\benvironmental\s*engineer',
        r'\baerospace\b', r'\bavionics\b',
        r'\bcad\b', r'\bdrafter\b', r'\bdrafting\b',
        r'\btechnician\b',  # broad catch for technical/trade roles
        r'\btest\s*engineer\b', r'\breliability\s*engineer\b',
        r'\blab\s*(research\s*)?engineer',
        r'\baquatic\s*design\b',
    ], None),

    # ── 8. Sales & Business Development ──────────────────────────────────────
    ("Sales & Business Development", [
        r'\bsales\b', r'\baccount\s*(executive|manager|rep|coordinator|specialist)\b',
        r'\bbusiness\s*develop', r'\bbdr\b', r'\bsdr\b',
        r'\breal\s*estate\s*(agent|broker|specialist)\b',
        r'\bclient\s*(relationship|success|service|develop|partner|engagement|manager)\b',
        r'\bcustomer\s*(success|service|support|experience|relation)\b',
        r'\b(inside|outside|retail|territory|regional)\s*sales\b',
        r'\baftermarket\s*sales\b',
        r'\bagent\b',
        r'\bbroker\b',
        r'\bcustomer\s*issue\b',
        r'\brep\b',
    ], None),

    # ── 9. Marketing & Communications ────────────────────────────────────────
    ("Marketing & Communications", [
        r'\bmarketing\b', r'\bbrand\b', r'\badvertis',
        r'\bpublic\s*relations\b', r'\bcommunications?\b',
        r'\bcontent\s*(writer|creator|strategist|manager|specialist|market|develop)\b',
        r'\bcopywriter\b', r'\bcopy\s*edit',
        r'\bsocial\s*media\b', r'\bmedia\s*(buyer|planner|coordinator|specialist|manager)\b',
        r'\bseo\b', r'\bsem\b', r'\bdigital\s*market',
        r'\becommerce\b', r'\be-?commerce\b',
        r'\bevent\s*(coordinator|planner|manager|specialist)\b',
        r'\bpr\s+(specialist|coordinator|manager|director)\b',
        r'\bjournalist\b', r'\breporter\b', r'\beditor\b', r'\beditorial\b',
        r'\bwriter\b', r'\bwriting\b',
        r'\bgraphic\s*design', r'\bcreative\s*(director|manager|design|specialist)\b',
        r'\bphotograph', r'\bvideograph', r'\bmultimedia\b',
        r'\bcampaign\b',
        r'\bfundraising\b', r'\bfund\s*rais',
        r'\bdonor\s*relat',
    ], None),

    # ── 10. Human Resources & Talent ─────────────────────────────────────────
    ("Human Resources", [
        r'\bhuman\s*resource', r'\bhr\s+(analyst|manager|specialist|coordinator|generalist|director|partner|business|intern|rep|assistant)\b',
        r'\bhr$', r'\b^hr\b',
        r'\brecruit', r'\btalent\s*(acqui|manage|develop|partner|special|coord)\b',
        r'\bstaffing\b', r'\bpayroll\b',
        r'\bpeople\s*(ops|operations|partner|manager|analyst)\b',
        r'\btraining\s*(specialist|coordinator|manager|develop|director)\b',
        r'\blearning\s*(&|and)\s*develop',
        r'\bdiversity\b', r'\binclusion\b',
        r'\bcompensation\b', r'\bbenefits\s*(analyst|specialist|coordinator|manager|admin)\b',
        r'\bemployee\s*(relation|engagement|experience)\b',
        r'\bonboarding\b',
        r'\borganizational\s*develop',
        r'\bhris\b',
    ], None),

    # ── 11. Consulting & Advisory ────────────────────────────────────────────
    ("Consulting & Advisory", [
        r'\bconsultant\b', r'\bconsulting\b',
        r'\badvisory\b', r'\badvisor\b(?!\s*(in\s*training))',
        r'\bstrateg(y|ic|ist)\b',
        r'\bmanagement\s*consult',
        r'\banalyst\b.*\bconsult',
        r'\bconsult.*\banalyst\b',
        r'\badvisory\s*(service|analyst)\b',
    ], None),

    # ── 12. Operations & Supply Chain ────────────────────────────────────────
    ("Operations & Supply Chain", [
        r'\boperations\b', r'\bsupply\s*chain\b', r'\blogistics\b',
        r'\bprocurement\b', r'\bpurchasing\b', r'\bbuyer\b',
        r'\bwarehouse\b', r'\bdistribution\b', r'\bshipping\b',
        r'\binventory\b', r'\bdemand\s*plan',
        r'\bproduction\s*(planner|manager|supervisor|coordinator|specialist)\b',
        r'\bplant\s*(manager|supervisor)\b',
        r'\bmanufacturing\s*(manager|supervisor|coordinator|specialist|assoc)\b',
        r'\bfacilities\b', r'\bmaintenance\s*(manager|tech|supervisor|engineer|planner)\b',
        r'\bquality\s*(control|assurance|manager|inspector|specialist|coordinator|technician)\b',
        r'\bcontinuous\s*improvement\b', r'\blean\b', r'\bsix\s*sigma\b',
        r'\bsafety\s*(manager|specialist|coordinator|officer|engineer|director)\b',
        r'\behs\b', r'\benvironment.*health.*safety\b',
    ], None),

    # ── 13. Research & Science ───────────────────────────────────────────────
    ("Research & Science", [
        r'\bresearch\b', r'\bscientist\b', r'\bscience\b',
        r'\bchemist\b', r'\bbiolog', r'\bphysicist\b',
        r'\blaboratory\b', r'\blab\s*(manager|director|assistant|coord|research)\b',
        r'\br\s*&\s*d\b', r'\brd&e\b',
        r'\bmicrobiol', r'\bneurosci', r'\bgeolog', r'\bbiosci',
        r'\bgenetic', r'\bbiochem', r'\btoxicol',
        r'\bfood\s*scien',
        r'\bpostdoc', r'\bresearch\s*(assistant|associate|fellow|analyst|coord)\b',
    ], None),

    # ── 14. Legal ────────────────────────────────────────────────────────────
    ("Legal", [
        r'\blegal\b', r'\battorney\b', r'\blawyer\b', r'\bjuris\b',
        r'\bparalegal\b', r'\blaw\s*clerk\b',
        r'\bcounsel\b(?!\s*or\b)',  # avoid "counselor"
        r'\blitigat', r'\bcontract\s*(special|analyst|manager|admin|coord)\b',
        r'\bregulatory\s*(affairs|specialist|analyst|compliance|manager)\b',
        r'\bip\s+(analyst|attorney|counsel|manager)\b',
        r'\bpatent\b',
    ], None),

    # ── 15. Government, Military & Public Service ────────────────────────────
    ("Government & Public Service", [
        r'\bgovernment\b', r'\bmilitary\b', r'\barmy\b', r'\bnavy\b',
        r'\bair\s*force\b', r'\bmarine', r'\bnational\s*guard\b',
        r'\bpolice\b', r'\bofficer\b(?=.*\b(police|law\s*enforcement|patrol|peace)\b)',
        r'\bfirefight', r'\blaw\s*enforcement\b',
        r'\bpublic\s*(policy|admin|affairs|servant|service)\b',
        r'\bcity\s*(planner|manager|admin|engineer)\b',
        r'\bpeace\s*corps\b', r'\bamericorps\b',
        r'\b2lt\b', r'\b2d\s*lt\b', r'\b1lt\b', r'\b1st\s*lt\b',
        r'\bcaptain\b', r'\blieutenant\b', r'\bsergeant\b',
        r'\bbandsman\b',
        r'\bintelligence\s*(analyst|officer|specialist)\b',
        r'\bpolicy\s*(analyst|advisor|specialist|manager)\b',
    ], None),

    # ── 16. Nonprofit & Social Services ──────────────────────────────────────
    ("Nonprofit & Social Services", [
        r'\bsocial\s*(work|service)\b', r'\bnonprofit\b', r'\bnon-?profit\b',
        r'\bvolunteer\s*(coord|manager|specialist)\b',
        r'\bcommunity\s*(organiz|outreach|develop|engagement|health|liaison)\b',
        r'\badvocacy\b', r'\badvocate\b',
        r'\byouth\s*(develop|coordinator|worker|program|specialist)\b',
        r'\bcase\s*manager\b',  # if not caught by healthcare
        r'\bhuman\s*service',
        r'\bfamily\s*service',
        r'\bchild\s*(welfare|protecti|life|develop)\b',
        r'\bdomestic\s*violen',
        r'\bshelter\b',
        r'\bpastoral\b', r'\bchaplain\b', r'\bminister\b', r'\bpastor\b',
        r'\breligious\b', r'\bfaith\b',
        r'\bvista\b',
        r'\bambassador\b',
    ], None),

    # ── 17. Design & Creative ────────────────────────────────────────────────
    ("Design & Creative", [
        r'\bdesign(?:er)?\b(?!.*\bengineer)',  # "Designer" but not "Design Engineer"
        r'\bgraphic\b', r'\billustrat', r'\bvisual\s*design',
        r'\binterior\s*design', r'\barchitect(?!ure)\b',
        r'\bux\b(?!.*engineer)', r'\buser\s*experience\b',
        r'\bart\s*(director|manager)\b',
        r'\banimator\b', r'\banimation\b',
        r'\bmusic', r'\bperform', r'\btheat',
    ], None),

    # ── 18. Management & Leadership (General) ────────────────────────────────
    ("Management & Leadership", [
        r'\b(vice\s*)?president\b', r'\bvp\b', r'\bceo\b', r'\bcoo\b', r'\bcto\b', r'\bcmo\b',
        r'\bchief\b', r'\bexecutive\s*director\b',
        r'\bgeneral\s*manager\b', r'\bstore\s*manager\b',
        r'\bmanager\b', r'\bdirector\b', r'\bsupervisor\b',
        r'\blead\b', r'\bhead\s*of\b',
        r'\bprogram\s*(manager|director|coordinator|specialist|officer)\b',
        r'\bproject\s*(manager|director|coordinator|lead)\b',
    ], None),

    # ── 19. Administrative & Support ─────────────────────────────────────────
    ("Administrative & Support", [
        r'\badmin', r'\bexecutive\s*assistant\b', r'\boffice\s*assist',
        r'\boffice\s*manager\b', r'\breceptionist\b', r'\bclerk\b',
        r'\bsecretar', r'\bfront\s*desk\b',
        r'\bcustomer\s*(service|support)\s*(rep|assoc|agent|specialist)\b',
        r'\bcall\s*center\b',
        r'\bdata\s*entry', r'\bfile\s*clerk',
    ], None),

    # ── 20. Construction & Trades ────────────────────────────────────────────
    ("Construction & Trades", [
        r'\bconstruction\b', r'\bcarpent', r'\belectrician\b',
        r'\bplumb', r'\bhvac\b', r'\bweld',
        r'\bforeman\b', r'\bsurvey', r'\bfield\s*(tech|worker|super)',
        r'\bsite\s*(super|manager|engineer)',
        r'\bauto\s*body\b', r'\bmechanic\b',
        r'\binstall(er|ation)\b',
        r'\btrade\b', r'\bapprentice\b',
        r'\bestimator\b', r'\binspector\b',
    ], None),

    # ── 21. Real Estate ──────────────────────────────────────────────────────
    ("Real Estate", [
        r'\breal\s*estate\b', r'\bproperty\s*(manag|analyst)\b',
        r'\bappraiser\b', r'\babstract',
        r'\bleasing\b', r'\brealtor\b',
    ], None),
]

# Fallback: use company industry to assign a default functional group
COMPANY_INDUSTRY_FALLBACK = {
    "1. Technology & Software": "Software & IT",
    "2. Healthcare & Medical": "Healthcare & Clinical",
    "3. Financial Services & Insurance": "Finance & Accounting",
    "4. Manufacturing & Industrial": "Operations & Supply Chain",
    "5. Education & Academia": "Education & Teaching",
    "6. Government & Public Service": "Government & Public Service",
    "7. Professional Services & Consulting": "Consulting & Advisory",
    "8. Retail & Consumer Products": "Sales & Business Development",
    "9. Energy & Utilities": "Engineering",
    "10. Construction, Engineering & Real Estate": "Engineering",
    "Other": "Other",
}


def classify_job(title, company, company_industry_group, specific_industry):
    """Classify a job title into a functional group."""
    if not title:
        return "Other"
    
    title_lower = title.lower().strip()
    company_lower = (company or "").lower().strip()
    specific_lower = (specific_industry or "").lower().strip()
    
    # Check each rule
    for category, title_patterns, company_patterns in CLASSIFICATION_RULES:
        for pattern in title_patterns:
            try:
                if re.search(pattern, title_lower):
                    # If company pattern is specified, must also match
                    if company_patterns:
                        for cp in company_patterns:
                            if re.search(cp, company_lower) or re.search(cp, specific_lower):
                                return category
                    else:
                        return category
            except re.error:
                continue
    
    # Fallback: generic titles -> use company industry group
    if company_industry_group and company_industry_group in COMPANY_INDUSTRY_FALLBACK:
        return COMPANY_INDUSTRY_FALLBACK[company_industry_group]
    
    return "Other"


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb['Full Data with Industries']
    
    # Add new column header
    new_col = 7  # Column G
    ws.cell(row=1, column=new_col, value="Job Title Classified Group")
    
    # Classify each row
    counts = Counter()
    examples = defaultdict(list)
    
    for row_idx in range(2, ws.max_row + 1):
        company = ws.cell(row=row_idx, column=1).value
        title = ws.cell(row=row_idx, column=3).value
        specific_industry = ws.cell(row=row_idx, column=4).value
        company_group = ws.cell(row=row_idx, column=5).value
        
        func_group = classify_job(title, company, company_group, specific_industry)
        ws.cell(row=row_idx, column=new_col, value=func_group)
        
        counts[func_group] += 1
        if len(examples[func_group]) < 5:
            examples[func_group].append(f"{title} @ {company}")
    
    # Print summary
    print("\n" + "=" * 80)
    print("JOB TITLE CLASSIFIED GROUPS SUMMARY")
    print("=" * 80)
    for group, count in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"\n  {group}: {count} records")
        for ex in examples[group]:
            print(f"    - {ex}")
    
    print(f"\nTotal: {sum(counts.values())} records across {len(counts)} groups")
    
    # Verify Ecolab/3M specifically
    print("\n" + "=" * 80)
    print("VERIFICATION - Ecolab roles:")
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Ecolab":
            title = ws.cell(row=row_idx, column=3).value
            old_group = ws.cell(row=row_idx, column=5).value
            new_group = ws.cell(row=row_idx, column=new_col).value
            print(f"  {title:55s} | OLD: {str(old_group):30s} | NEW: {new_group}")
    
    print("\nVERIFICATION - 3M roles:")
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "3M":
            title = ws.cell(row=row_idx, column=3).value
            old_group = ws.cell(row=row_idx, column=5).value
            new_group = ws.cell(row=row_idx, column=new_col).value
            print(f"  {title:55s} | OLD: {str(old_group):30s} | NEW: {new_group}")
    
    # Save
    print(f"\nSaving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print("Done!")


if __name__ == "__main__":
    main()
